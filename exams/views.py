from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.db.models import Sum, Avg, Count, Min, Max, Q
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST, require_http_methods
import json, os, io, datetime
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from django.db import transaction

# ✅ Always use get_user_model
from django.contrib.auth import get_user_model
User = get_user_model()
# Do NOT assign settings.AUTH_USER_MODEL (that's a string) — use the model loaded earlier


# Your models
from .models import Quiz, Question, Choice, StudentQuizAttempt, ActionLog, Answer, Class, Subject,RetakeRequest
from users.models import Notification
from .utils import log_action
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Avg





def home(request):
    return render(request, "exams/home.html")



# --------------------------------------------------------------#

def is_teacher(user):
    return user.is_authenticated and user.role == 'teacher'

def is_admin(user):
    return user.is_authenticated and user.role in ('admin', 'superadmin')

def is_teacher_or_admin(user):
    return user.is_authenticated and user.role in ('teacher', 'admin', 'superadmin')

def is_admin_or_superadmin(user):
    return user.is_authenticated and user.role in ["admin", "superadmin"]


@login_required
@user_passes_test(is_admin_or_superadmin)
def superadmin_dashboard(request):
    if request.user.role !=  "superadmin":
        return HttpResponseForbidden("Unauthorized")
    # counts
    approved = User.objects.filter(role__in=['admin', 'teacher', 'student'], approved=True).count()
    pended = User.objects.filter(approved=False).count()
    teachers = User.objects.filter(role='teacher', approved=True).count()
    students = User.objects.filter(role='student', approved=True).count()

    # leaderboard: top students by average total_score (only graded attempts)
    leaderboard = (
        StudentQuizAttempt.objects.filter(is_submitted=True)
        .values("student__username")
        .annotate(avg_score=Avg("score"))
        .order_by("-avg_score")[:10]  # top 10
    )

    # recent action logs
    actions = ActionLog.objects.order_by('-timestamp')[:20]

    notifications = Notification.objects.filter(role='superadmin').order_by('-created_at')[:10]
    

    context = {
        "approved": approved,
        "pended": pended,
        "teachers": teachers,
        "students": students,
        "leaderboard": leaderboard,
        "actions": actions,
        "notifications": notifications,
        "total_admins": User.objects.filter(role="admin", approved=True).count(),
        "total_quizzes": Quiz.objects.count(),
        "total_classes": Class.objects.count(),
        'total_subjects': Subject.objects.count(),
    }
    return render(request, "exams/superadmin_dashboard.html", context)




# helper: admin check
def is_admin(user):
    return user.is_authenticated and user.role in ("admin", "superadmin")

# Admin dashboard page (HTML shell)
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    # Render the shell; the dynamic data comes from admin_dashboard_data
    return render(request, "exams/admin_dashboard.html", {})


# Admin dashboard data endpoint (GET -> fetch data; POST -> perform actions such as approve/reject/broadcast/download)
@login_required
def admin_dashboard_data(request):
    if request.user.role not in ("admin", "superadmin"):
        return JsonResponse({"error": "forbidden"}, status=403)

    # Handle POST actions: approve/reject/pending user, broadcast, download, etc
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode())
        except Exception:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        action_type = payload.get("action_type")

        # ---------------------------
        # Approve / Reject / Pend user
        # ---------------------------
        if action_type == "update_user_status":
            if request.user.role not in ("admin", "superadmin"):
                return JsonResponse({"error": "forbidden"}, status=403)
            user_id = payload.get("user_id")
            new_status = payload.get("status")  # "approve"/"reject"/"pending"
            target_user = get_object_or_404(User, id=user_id)
            old_approved = target_user.approved
            if new_status == "approve":
                target_user.approved = True
            elif new_status == "reject":
                # we don't delete here automatically; mark approved False and optionally flag
                target_user.approved = False
            elif new_status == "pending":
                target_user.approved = False
            else:
                return JsonResponse({"error": "invalid status"}, status=400)
            target_user.save(update_fields=["approved"])
            # Log action
            ActionLog.objects.create(
                user=request.user,
                action_type=f"Updated user status -> {new_status}",
                model_name="User",
                object_id=str(target_user.id),
                details={"old_approved": old_approved, "new_approved": target_user.approved},
            )
            return JsonResponse({"ok": True, "message": f"{target_user.username} set to {new_status}"})

        # ---------------------------
        # Broadcast (admin can to teachers/students; teacher can to students only)
        # ---------------------------
        if action_type == "broadcast":
            role = payload.get("role")  # 'teacher' or 'student'
            message = payload.get("message", "").strip()
            if not role or not message:
                return JsonResponse({"error": "role & message required"}, status=400)
            # permission check: teacher can only send to 'student'
            if request.user.role == "teacher" and role != "student":
                return JsonResponse({"error": "forbidden"}, status=403)
            # admins can send to both teachers/students
            recipients = User.objects.filter(role=role, approved=True)
            created = 0
            for r in recipients:
                Notification.objects.create(sender=request.user, recipient=r, message=message, role=role, is_broadcast=True)
                created += 1
                ActionLog.objects.create(
                user=request.user,
                action_type="Broadcast",
                model_name="Notification",
                object_id="bulk",
                details={"role": role, "count": created, "message": message[:200]},
            )
            return JsonResponse({"ok": True, "message": f"Broadcast sent to {created} {role}(s).", "count": created})

        # ---------------------------
        # Download all attempts (excel or pdf)
        # payload: {action: "download", format: "pdf"|"excel"}
        # ---------------------------
        if action_type == "download":
            if request.user.role not in ("admin", "superadmin"):
                return JsonResponse({"error": "forbidden"}, status=403)

            fmt = payload.get("format", "pdf")
            # Query attempts (all submitted attempts)
            attempts_qs = StudentQuizAttempt.objects.filter(is_submitted=True).select_related("quiz", "student")

            # Build in-memory file
            if fmt == "excel":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "AllAttempts"
                headers = ["Student", "Username", "Quiz", "Score", "Started At", "submitted At"]
                ws.append(headers)
                for a in attempts_qs:
                    student_username = a.student.username if hasattr(a.student, "username") else str(a.student)
                    started = a.started_at.strftime("%Y-%m-%d %H:%M") if a.started_at else ""
                    submitted_at = a.submitted_at.strftime("%Y-%m-%d %H:%M") if getattr(a, "submitted_at", None) else ""
                    ws.append([getattr(a.student, "get_full_name", student_username) or student_username, student_username, a.quiz.title, a.score, started, submitted_at])
                # save to bytes
                bio = io.BytesIO()
                wb.save(bio)
                bio.seek(0)
                ActionLog.objects.create(user=request.user, action_type="Downloaded all attempts (excel)", model_name="StudentQuizAttempt", object_id="all", details={"count": attempts_qs.count()})
                response = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response["Content-Disposition"] = "attachment; filename=all_attempts.xlsx"
                return response

            # PDF via ReportLab
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            elements.append(Paragraph("All Attempts Results", styles["Title"]))
            elements.append(Spacer(1, 12))

            table_data = [["Student", "Username", "Exam", "Score", "Started At", "Completed At"]]
            for a in attempts_qs:
                student_username = a.student.username if hasattr(a.student, "username") else str(a.student)
                started = a.started_at.strftime("%d-%m-%Y %H:%M") if a.started_at else ""
                submitted_at = a.submitted_at.strftime("%d-%m-%Y %H:%M") if getattr(a, "submitted_at", None) else ""
                table_data.append([getattr(a.student, "get_full_name", student_username) or student_username, student_username, a.quiz.title, str(a.score), started, submitted_at])
            tbl = Table(table_data, repeatRows=1)
            tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.gray),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),]))
            elements.append(tbl)
            doc.build(elements)
            buffer.seek(0)
            ActionLog.objects.create(user=request.user, action_type="Downloaded all attempts (pdf)", model_name="StudentQuizAttempt", object_id="all", details={"count": attempts_qs.count()})
            response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = "attachment; filename=all_attempts.pdf"
            return response

        return JsonResponse({"error": "unknown action"}, status=400)

    # -----------------------
    # GET: return dashboard data JSON
    # Supports pagination parameters: logs_page, quizzes_page, quizzes_page_size, logs_page_size
    # -----------------------
    # Stats
    stats = {
        "total_users": User.objects.filter(role__in=["admin", "teacher", "student"], approved=True).count(),
        "admins": User.objects.filter(role="admin", approved=True).count(),
        "teachers": User.objects.filter(role="teacher", approved=True).count(),
        "students": User.objects.filter(role="student", approved=True).count(),
        "pending_users": User.objects.filter(approved=False).count(),
        "classes": Class.objects.count(),
        "subjects": Subject.objects.count(),
        "quizzes": Quiz.objects.count(),
    }
    
    # Pending users (basic)
    pending_list_qs = User.objects.filter(approved=False).order_by("-date_joined")
    pending_list = list(pending_list_qs.values("id", "username", "email", "role", "date_joined")[:50])

    # Action logs paginated
    logs_page = int(request.GET.get("logs_page", 1))
    logs_page_size = int(request.GET.get("logs_page_size", 5))
    logs_qs = ActionLog.objects.order_by("-timestamp")
    paginator_logs = Paginator(logs_qs, logs_page_size)
    page_logs = paginator_logs.get_page(logs_page)
    logs = [
        {"action_type": l.action_type, "user": (l.user.username if l.user else "system"), "timestamp": l.timestamp.isoformat(), "details": l.details}
        for l in page_logs
    ]

    # Leaderboard (top students by average obtained_marks across answers)
    leaderboard_qs = StudentQuizAttempt.objects.filter(is_submitted=True).values("student__username", 'student__first_name', 'student__last_name', 'student__student_class__name').annotate(avg_score=Avg("score")).order_by("-avg_score")[:10]

    leaderboard = [{"username": x["student__username"], "first_name": x["student__first_name"], "last_name": x['student__last_name'], "class": x["student__student_class__name"], "avg_score": float(x["avg_score"] or 0)} for x in leaderboard_qs]

    # Class performance (avg obtained marks grouped by class)
    class_perf_qs = Answer.objects.values("attempt__quiz__subject__school_class__name").annotate(avg_score=Avg("obtained_marks")).order_by("attempt__quiz__subject__school_class__name")
    class_performance = [
        {"class_name": row["attempt__quiz__subject__school_class__name"] or "Unknown", "avg_score": float(row["avg_score"] or 0)}
        for row in class_perf_qs
    ]

    # Available quizzes (paginated) - show basic metadata
    quizzes_page = int(request.GET.get("quizzes_page", 1))
    quizzes_page_size = int(request.GET.get("quizzes_page_size", 10))
    quizzes_qs = Quiz.objects.select_related("subject", "created_by").order_by("-created_at")
    paginator_quiz = Paginator(quizzes_qs, quizzes_page_size)
    page_quiz = paginator_quiz.get_page(quizzes_page)
    quizzes = [
        {
            "id": q.id,
            "title": q.title,
            "subject": q.subject.name,
            "class_name": q.subject.school_class.name,
            "created_by": getattr(q.created_by, "username", str(q.created_by)),
            "start_time": q.start_time.isoformat(),
            "end_time": q.end_time.isoformat(),
            "is_published": q.is_published,
            "allow_retake": getattr(q, "allow_retake", False),
        }
        for q in page_quiz
    ]

    # Notifications for this admin (last 10)
    notifications_qs = Notification.objects.filter(recipient=request.user, is_read=False).order_by("-created_at")[:5]
    notifications = [{"id": n.id, "message": n.message, "sender": getattr(n.sender, "username", None), "created_at": n.created_at.isoformat()} for n in notifications_qs]
   

    return JsonResponse({
        "stats": stats,
        "pending_list": pending_list,
        "logs": logs,
        "logs_total_pages": paginator_logs.num_pages,
        "leaderboard": leaderboard,
        "class_performance": class_performance,
        "notifications": notifications,
        "quizzes": quizzes,
        "quizzes_total_pages": paginator_quiz.num_pages,
    })

#-------------------------- Teacher Dashboard --------------------------###

def is_teacher(user):
    return user.is_authenticated and user.role == "teacher"



# ----------------- TEACHER DASHBOARD -------------------



def is_teacher(user):
    return user.is_authenticated and user.role == "teacher"


# ============================
# Teacher Dashboard
# ============================
@login_required
@user_passes_test(is_teacher_or_admin)
def teacher_dashboard(request):
    """Render teacher dashboard page"""
    return render(request, "teacher_dashboard.html")


@login_required
@user_passes_test(is_teacher_or_admin)
def teacher_dashboard_data(request):
    """Provide teacher dashboard data (JSON) with pagination"""
    teacher = request.user

    student_class = teacher.student_class   


    if teacher.is_authenticated and teacher.role in ["teacher", "admin", "superadmin"]:
        teacher = User.objects.filter(id=teacher.id).first()
        student = User.objects.filter(role="student", student_class=student_class).first()

        # Classes linked to teacher
        classes = Quiz.objects.filter(created_by=teacher).values("subject__school_class").distinct()

        # Quizzes created by teacher
        quizzes = Quiz.objects.filter(subject__school_class=student_class).annotate(
            attempt_count=Count("studentquizattempt")).order_by("-created_at")

        quiz_page = Paginator(quizzes, 5).get_page(request.GET.get("quiz_page", 1))
        
        # Performance stats
        performance = (
            Answer.objects.filter(attempt__student__student_class=student_class)
            .values("attempt__student__username")
            .annotate(avg_score=Avg("obtained_marks"))
        )

        # Notifications
        notifications = Notification.objects.filter(recipient=teacher).order_by("-created_at")
        notif_page = Paginator(notifications, 5).get_page(request.GET.get("notif_page", 1))

        # Logs
        logs = ActionLog.objects.filter(user=teacher).order_by("-timestamp")
        logs_page = Paginator(logs, 5).get_page(request.GET.get("log_page", 1))

        # Pending grading (subjective answers without marks)
        pending_answers = Answer.objects.filter(
            attempt__quiz__created_by=teacher, obtained_marks=0.0, text_answer__isnull=False
        ).order_by("-graded_at")
        grade_page = Paginator(pending_answers, 5).get_page(request.GET.get("grade_page", 1))


        # ---------- SUMMARY ----------
        # my_quizzes (same as total_quizzes but keep name used in some templates)
        my_quizzes = Quiz.objects.filter(created_by=request.user).count()

        # Objectives count = number of objective questions in this teacher's quizzes
        objectives = Question.objects.filter(quiz__created_by=teacher, question_type="objective").count()

        # Graded subjectives count (answers that are subjective and graded)
        graded_subjectives = Answer.objects.filter(
            question__quiz__created_by=teacher,
            question__question_type="subjective",
            is_pending=False  
        ).count() # graded

        # Pending subjectives (need teacher grading)
        pending_subjectives = Answer.objects.filter(
            question__quiz__created_by=teacher,
            question__question_type="subjective",
            is_pending=True
        ).count()
        attempts = StudentQuizAttempt.objects.filter(student=student ).count()

    data = {
        "summary": {
            "total_quizzes": quizzes.count(),
            "my_quizzes": my_quizzes,
            "objectives": objectives,
            "graded": graded_subjectives,
            "pended": pending_subjectives,
            "student_class": str(student_class),
            "teacher" : str(teacher),
            "attempts": str(attempts),
            "quiz_page": str(quiz_page),
            "notif_page": str(notif_page),
            "logs_page": str(logs_page),
            "grade_page": str(grade_page),
            "total_students": User.objects.filter(student_class=request.user.student_class, role="student").count(),
        },
        "quizzes": [
            {"id": q.id, "title": q.title, "attempts": q.attempt_count, "subject": q.subject.name, "class_name": q.subject.school_class.name, "created_at": q.created_at.isoformat(), "start_time": q.start_time.isoformat(), "end_time": q.end_time.isoformat(), "is_published": q.is_published, "allow_retake": getattr(q, "allow_retake", False)}
            for q in quiz_page
        ],
        "performance": list(performance),
        "notifications": [
            {
                "id": n.id,
                "message": n.message,
                "is_read": n.is_read,
                "created_at": n.created_at.isoformat()
            }
            for n in notif_page
        ],
        "logs": [
            {
                "action_type": l.action_type,
                "user": (l.user.username if l.user else "system"),
                "timestamp": l.timestamp.isoformat(),
                "details": l.details,
            }
            for l in logs_page
        ],
        "grading": [
            {
                "id": a.id,
                "student": a.attempt.student.username,
                "quiz": a.attempt.quiz.title,
                "question": a.question.text,
                "answer": a.text_answer,
                "marks": a.obtained_marks,
            }
            for a in grade_page
        ],
        "pagination": {
            "quiz_num_pages": quiz_page.paginator.num_pages,
            "notif_num_pages": notif_page.paginator.num_pages,
            "logs_num_pages": logs_page.paginator.num_pages,
            "grade_num_pages": grade_page.paginator.num_pages,
        }
    }
   
    return JsonResponse(data)


# ============================
# Teacher Broadcast
# ============================
@login_required
@user_passes_test(is_teacher)
def teacher_broadcast(request):
    """Teacher sends a broadcast"""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    message = request.POST.get("message")
    audience = request.POST.get("audience")  # "students" | "admin"

    if not message or not audience:
        return JsonResponse({"error": "Missing fields"}, status=400)

    if audience == "students":
        recipients = User.objects.filter(role="student")
    elif audience == "admin":
        recipients = User.objects.filter(role="admin")
    else:
        return JsonResponse({"error": "Invalid audience"}, status=400)

    for r in recipients:
        Notification.objects.create(recipient=r, message=message)

    ActionLog.objects.create(
        user=request.user,
        action_type="Teacher Broadcast",
        model_name="Notification",
        object_id=str(request.user.id),
        details={"message": message, "audience": audience},
    )

    return JsonResponse({"success": True, "message": f"Broadcast sent to {audience}."})


# ============================
# Grading Endpoint
# ============================
@login_required
@user_passes_test(is_teacher)
def grade_answer(request, answer_id):
    """Grade a student's subjective answer"""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    answer = get_object_or_404(Answer, id=answer_id)
    marks = float(request.POST.get("marks", 0))
    feedback = request.POST.get("feedback", "")

    answer.obtained_marks = marks
    answer.feedback = feedback
    answer.graded_by = request.user
    answer.graded_at = timezone.now()
    answer.is_pending = False
    answer.save()

    # Notify student
    Notification.objects.create(
        recipient=answer.attempt.student,
        message=f"Your answer for '{answer.question.text}' was graded: {marks} marks."
    )

    # Log grading
    ActionLog.objects.create(
        user=request.user,
        action_type="Graded Answer",
        model_name="Answer",
        object_id=str(answer.id),
        details={
            "student": answer.attempt.student.username,
            "quiz": answer.attempt.quiz.title,
            "marks": marks,
        },
    )

    return JsonResponse({"success": True, "message": "Answer graded."})

@login_required
@require_POST
def mark_notification_read(request, notif_id):
    notif = get_object_or_404(Notification, id=notif_id, user=request.user)
    notif.is_read = True
    notif.save()

    return JsonResponse({
        "success": True,
        "notif_id": notif.id,
        "message": "Notification marked as read."
    })



# ----------------- STUDENT REVIEW  on teacher dashboard-------------------
@login_required
@user_passes_test(is_teacher)
def student_review(request, student_id):
    student = get_object_or_404(User, id=student_id, role="student")
    attempts = StudentQuizAttempt.objects.filter(student=student).select_related("quiz")
    return render(request, "teacher/student_review.html", {"student": student, "attempts": attempts})



# ----------------- RETAKE REQUEST -------------------
@login_required
@user_passes_test(is_teacher)
def approve_retake(request, quiz_id, student_id):
    student = get_object_or_404(User, id=student_id, role="student")
    quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)

    attempt, _ = StudentQuizAttempt.objects.get_or_create(student=student, quiz=quiz)
    attempt.retake_allowed = True
    attempt.is_submitted = False
    attempt.end_time = None
    attempt.retake_count += 1
    attempt.save()
  
    Notification.objects.create(user=student, message=f"You can now retake Exam: {quiz.title}")
    ActionLog.objects.create(user=request.user, action="Approved Retake", model_name="Exam", object_id=str(quiz.id))

    return JsonResponse({"success": True})

# ----------------- BROADCAST MESSAGE -------------------
@login_required
@user_passes_test(is_teacher)
def broadcast_message(request):
    if request.method == "POST":
        message = request.POST.get("message")
        students = User.objects.filter(role="student", school_class__in=request.user.classes.all())
        for student in students:
            Notification.objects.create(recipient=student, message=message, sender=request.user, is_broadcast=True, defaults={"role": "student", "created_at": timezone.now()})

        ActionLog.objects.create(user=request.user, action="Broadcast Message", model_name="BroadcastMessage", object_id=str(request.user.id))
        return JsonResponse({"success": True})
    return JsonResponse({"error": "Invalid request"}, status=400)



def teacher_broadcast(request):
    """Teacher sends broadcast notifications to admins, students, or all."""
    if request.method == "POST":
        message = request.POST.get("message")
        target = request.POST.get("target")  # admins | students | all

        if not message or not target:
            return JsonResponse({"ok": False, "error": "Message and target required."})

        # recipients
        if target == "admins":
            recipients = User.objects.filter(role="admin")
        elif target == "students":
            recipients = User.objects.filter(role="student", student_class__in=request.user.classroom_set.all())
        else:  # all
            recipients = User.objects.filter(role__in=["admin", "student"])

        # create notifications
        notifications = [
            Notification(
                user=r,
                message=f"📢 {message}",
                created_at=timezone.now(),
                is_read=False,
            )
            for r in recipients
        ]
        Notification.objects.bulk_create(notifications)

        # log action
        log_action(request.user, f"Broadcasted to {target}: {message}")

        return JsonResponse({"ok": True, "message": f"Broadcast sent to {target}"})

    return JsonResponse({"ok": False, "error": "Invalid request"})




# ----------------- DOWNLOAD REPORTS -------------------
@login_required
@user_passes_test(is_teacher)
def download_student_report(request, student_id):
    student = get_object_or_404(User, id=student_id, role="student")
    attempts = StudentQuizAttempt.objects.filter(student=student)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    story = []
    styles = getSampleStyleSheet()

    story.append(Paragraph(f"<b>School Name</b>", styles["Title"]))
    story.append(Paragraph(f"Student: {student.get_full_name()}", styles["Heading2"]))
    story.append(Spacer(1, 12))

    for attempt in attempts:
        story.append(Paragraph(f"Quiz: {attempt.quiz.title} | Date: {attempt.start_time}", styles["Normal"]))
        for ans in attempt.answer.all():
            story.append(Paragraph(f"- {ans.question.text}: {ans.obtained_marks}", styles["Normal"]))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{student.username}_report.pdf"'
    return response

@login_required
@user_passes_test(is_teacher)
def download_quiz_report(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    story = []
    styles = getSampleStyleSheet()

    story.append(Paragraph(f"<b>School Name</b>", styles["Title"]))
    story.append(Paragraph(f"Quiz Report: {quiz.title}", styles["Heading2"]))
    story.append(Spacer(1, 12))

    for attempt in attempts:
        student = attempt.student
        story.append(Paragraph(f"Student: {student.get_full_name()} | Date: {attempt.start_time}", styles["Normal"]))
        for ans in attempt.answer.all():
            story.append(Paragraph(f"- {ans.question.text}: {ans.obtained_marks}", styles["Normal"]))
        story.append(Spacer(1, 6))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="quiz_{quiz.id}_report.pdf"'
    return response

# ----------------- MARK NOTIFICATION READ -------------------
@login_required
def mark_notification_read(request, notification_id):
    notif = get_object_or_404(Notification, id=notification_id, user=request.user)
    notif.is_read = True
    notif.save()
    return JsonResponse({"success": True})


# ----------------------Teacher dashboard data endpoint (JSON) (for AJAX refresh)-------------------------#



###--------------------------------Student Dashboard -------------------------------###

def is_student(user):
    return user.is_authenticated and user.role == 'student' 



@login_required
@user_passes_test(is_student)
def student_dashboard(request):
    """Render the student dashboard shell. Data comes from student_dashboard_data (AJAX)."""
    return render(request, "exams/student_dashboard.html")


@login_required
@user_passes_test(is_student)
def student_dashboard_data(request):
    """
    Returns JSON containing:
      - notifications (paginated)
      - summary (counts)
      - available_quizzes (paginated)
      - past_attempts (paginated)
      - leaderboard (top 10)
      - performance_chart (subject -> %)
    Query params (optional):
      notif_page, notif_page_size, quizzes_page, quizzes_page_size, attempts_page, attempts_page_size
    """
    student = request.user
    student_class = getattr(student, "student_class", None)

    # pagination params
    notif_page = int(request.GET.get("notif_page", 1))
    notif_page_size = int(request.GET.get("notif_page_size", 6))
    quizzes_page = int(request.GET.get("quizzes_page", 1))
    quizzes_page_size = int(request.GET.get("quizzes_page_size", 6))
    attempts_page = int(request.GET.get("attempts_page", 1))
    attempts_page_size = int(request.GET.get("attempts_page_size", 6))

    # ---------------- Notifications (paginated) ----------------
    notif_qs = Notification.objects.filter(recipient=student).order_by("-created_at")
    notif_p = Paginator(notif_qs, notif_page_size)
    notif_page_obj = notif_p.get_page(notif_page)
    notifications = [
        {"id": n.id, "message": n.message, "is_read": n.is_read, "created_at": n.created_at.isoformat()}
        for n in notif_page_obj
    ]
    notif_meta = {"page": notif_page_obj.number, "pages": notif_p.num_pages, "total": notif_p.count}

    # ---------------- Summary ----------------
    total_attempts = StudentQuizAttempt.objects.filter(student=student).count()
    auto_graded_count = Answer.objects.filter(attempt__student=student, is_pending=False).count()
    pending_subjectives = Answer.objects.filter(attempt__student=student, is_pending=True).count()

    summary = {
        "total_attempts": total_attempts,
        "auto_graded_count": auto_graded_count,
        "pending_subjectives": pending_subjectives,
    } 

    # ---------------- Available quizzes (paginated) ----------------

    now = timezone.now()
    if student_class is None:
        available_qs = Quiz.objects.none()
    else:
        available_qs = Quiz.objects.filter(
            school_class__name=student_class,
            is_published=True
        ).order_by("-created_at")

        # exclude quizzes already completed without retake allowed
        exclude_ids = StudentQuizAttempt.objects.filter(
            student=student, is_submitted=True, retake_allowed=False,
        ).values_list("quiz_id", flat=True)
        available_qs = available_qs.exclude(id__in=exclude_ids) 

    qp = Paginator(available_qs, quizzes_page_size)
    qp_obj = qp.get_page(quizzes_page)

    quizzes_data = []
    # build rich data (can't easily get allow_retake from values; use getattr)
    for q in qp_obj:
        # check last attempt status
        last_attempt = StudentQuizAttempt.objects.filter(student=student, quiz=q).order_by("-started_at").first()
        already_submitted = StudentQuizAttempt.objects.filter(student=student, quiz=q, is_submitted=True).exists()
        student_retake_override = bool(last_attempt and getattr(last_attempt, "retake_allowed", False))
        allow_retake_global = getattr(q, "allow_retake", False) if hasattr(q, "allow_retake") else False
        # Added lately for great UI/UX. 
        latest_request = RetakeRequest.objects.filter(student=request.user, quiz=q).last()
        retake_status = latest_request.status if latest_request else None

        quizzes_data.append({
            "id": q.id,
            "title": q.title,
            "subject": q.subject.name if q.subject else "",
            "class_name": q.subject.school_class.name if (q.subject and q.subject.school_class) else "",
            "start_time": q.start_time.isoformat() if q.start_time else None,
            "end_time": q.end_time.isoformat() if q.end_time else None,
            "duration_minutes": getattr(q, "duration_minutes", None),
            "is_published": bool(q.is_published),
            "allow_retake": bool(allow_retake_global),
            "already_submitted": bool(already_submitted),
            "student_retake_override": student_retake_override,
            "retake_status": retake_status,  # 🔹 added
        })
            
    quizzes_meta = {"page": qp_obj.number, "pages": qp.num_pages, "total": qp.count}

    # ---------------- Past attempts (paginated) ----------------
    attempts_qs = StudentQuizAttempt.objects.filter(student=student).select_related("quiz").order_by("-started_at")
    ap = Paginator(attempts_qs, attempts_page_size)
    ap_obj = ap.get_page(attempts_page)

    past_attempts = []
    for a in ap_obj:
        # compute totals for this attempt using Answer model
        answer = Answer.objects.filter(attempt=a).select_related("question", "selected_choice")
        total_marks = 0
        obtained = 0.0
        pending_subjectives_count = 0
        wrong_review = []
        for ans in answer:
            q_marks = getattr(ans.question, "marks", 0) or 0
            total_marks += q_marks
            if ans.is_pending:
                pending_subjectives_count += 1
                # subjective might not have marks yet; don't include
            else:
                # for objective and graded subjective
                obtained += float(ans.obtained_marks or 0.0)
            # for review: detect wrong objective answer
            # if question is objective and selected_choice exists but is incorrect
            if getattr(ans.question, "question_type", getattr(ans.question, "type", None)) in ("objective", "multiple_choice"):
                if not (ans.selected_choice and getattr(ans.selected_choice, "is_correct", False)):
                    # get correct answer text(s)
                    correct_choices = ans.question.choices.filter(is_correct=True).values_list("text", flat=True)
                    correct_text = ", ".join(correct_choices) if correct_choices else ""
                    user_ans = ans.selected_choice.text if ans.selected_choice else "-"
                    wrong_review.append({
                        "question": ans.question.text,
                        "your_answer": user_ans,
                        "correct_answer": correct_text
                    })

        past_attempts.append({
            "attempt_id": a.id,
            "quiz": a.quiz.title if a.quiz else "",
            "score": float(a.score or obtained),
            "obtained_marks": obtained,
            "total_marks": total_marks,
            "pending_subjectives": pending_subjectives_count,
            "is_submitted": bool(a.is_submitted),
            "started_at": a.started_at.isoformat() if getattr(a, "started_at", None) else None,
            "submitted_at": getattr(a, "submitted_at", None).isoformat() if getattr(a, "submitted_at", None) else None,
            "wrong_answer": wrong_review,
            "retake_count": getattr(a, "retake_count", 0),
        })

    attempts_meta = {"page": ap_obj.number, "pages": ap.num_pages, "total": ap.count}

    # ---------------- Leaderboard (top 10 students) ----------------
    lb_qs = (
        StudentQuizAttempt.objects.filter(is_submitted=True)
        .values("student__id", "student__username", "student__first_name", "student__last_name")
        .annotate(avg_score=Avg("score"))
        .order_by("-avg_score")[:10]
    )
    leaderboard = [
        {
            "student_id": r["student__id"],
            "username": r.get("student__username"),
            "full_name": f"{r.get('student__first_name') or ''} {r.get('student__last_name') or ''}".strip(),
            "avg_score": float(r["avg_score"] or 0)
        }
        for r in lb_qs
    ]

    # ---------------- Performance Chart (per subject in student's class) ----------------

    perf_qs = Answer.objects.filter(
        attempt__student=student,
        attempt__quiz__subject__school_class=student_class
    ).values("attempt__quiz__subject__name").annotate(
        obtained=Sum("obtained_marks"),
        possible=Sum("question__marks")
    )

    performance_chart = []
    for r in perf_qs:
        subj = r.get("attempt__quiz__subject__name") or "Unknown"
        obtained = float(r.get("obtained") or 0)
        possible = float(r.get("possible") or 0) or 0.0
        pct = round((obtained / possible) * 100, 2) if possible > 0 else 0.0
        performance_chart.append({"subject": subj, "obtained": obtained, "possible": possible, "percentage": pct})


    # ---------------- Return JSON ----------------

    return JsonResponse({
        "notifications": notifications,
        "notifications_meta": notif_meta,
        "summary": summary,
        "available_quizzes": quizzes_data,
        "available_quizzes_meta": quizzes_meta,
        "past_attempts": past_attempts,
        "past_attempts_meta": attempts_meta,
        "leaderboard": leaderboard,
        "performance_chart": performance_chart,
    })


@require_POST
@login_required
@user_passes_test(is_student)
def api_notifications_mark_read(request):
    """Mark notification read (AJAX POST: {id: notification_id})"""
    import json
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid json"}, status=400)

    nid = payload.get("id")
    if not nid:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    notif = get_object_or_404(Notification, id=nid, recipient=request.user)
    notif.is_read = True
    notif.save(update_fields=["is_read"])
    return JsonResponse({"ok": True, "id": nid})


### ------------------------------ Student_dashboard Ended ----------------------------------------------------##

def _is_admin(user):
    return user.is_authenticated and user.role in ("admin", "superadmin")

def _is_teacher(user):
    return user.is_authenticated and user.role == "teacher"


# Broadcast endpoint
# -------------------
@login_required
@require_POST
def api_broadcast(request):
    """
    POST JSON { "role": "student"|"teacher", "message": "..." }
    - Admins (admin/superadmin) may broadcast to 'teacher' or 'student'
    - Teachers may broadcast only to 'student'
    Creates Notification objects for each recipient and logs ActionLog.
    """
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    role = payload.get("role")
    message = (payload.get("message") or "").strip()
    if not role or not message:
        return JsonResponse({"ok": False, "error": "role and message required"}, status=400)

    # Permissions: teacher can only send to students
    if request.user.role == "teacher" and role != "student":
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # Only allow admin/teacher to broadcast
    if request.user.role not in ("admin", "superadmin", "teacher"):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    recipients = User.objects.filter(role=role, approved=True)
    created = 0
    for r in recipients:
        # Notification.objects.create(sender=request.user, recipient=r, role=role, message=message, created_at=timezone.now())
        created += 1

    ActionLog.objects.create(
        user=request.user,
        action="Broadcast sent",
        model_name="Notification",
        object_id="bulk",
        details={"role": role, "count": created, "sample": message[:120]},
        timestamp=timezone.now()
    )

    return JsonResponse({"ok": True, "message": f"Broadcast sent to {created} {role}(s).", "count": created})


# -------------------
# Get unread notifications for current user
# -------------------
@login_required
def api_notifications_unread(request):
    """
    GET -> returns unread (is_read=False) notifications for request.user
    """
    qs = Notification.objects.filter(recipient=request.user, is_read=False).order_by("-created_at")
    data = [
        {
            "id": n.id,
            "message": n.message,
            "sender": n.sender.username if n.sender else None,
            "created_at": n.created_at.isoformat()
        }
        for n in qs
    ]
    return JsonResponse({"ok": True, "notifications": data})


# -------------------
# Mark notification as read
# -------------------

@login_required
@require_POST
def api_notifications_mark_read(request):
    """
    # POST JSON { "id": <notification_id> } -> set is_read = True for that notification if it belongs to current user
    """
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    nid = payload.get("id")
    if not nid:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    notif = get_object_or_404(Notification, id=nid, recipient=request.user)
    notif.is_read = True
    notif.save(update_fields=["is_read"])

    ActionLog.objects.create(
        user=request.user,
        action="Read notification",
        model_name="Notification",
        object_id=str(nid),
        # details={"message_sample": (notif.message[:80] if notif.message else "")},
        timestamp=timezone.now()
    )

    return JsonResponse({"ok": True, "message": "marked read", "id": nid})


####--------------------------No 1 Endpoint Ended----------------------------####


@require_http_methods(["GET"])
def api_notifications_unread(request):
    """Return unread notifications for current user."""
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    from users.models import Notification  # import here to avoid circular issues
    qs = Notification.objects.filter(recipient=request.user, is_read=False).order_by("-created_at")
    notifs = [{
        "id": n.id,
        "message": n.message,
        "sender": n.sender.username if n.sender else None,
        "created_at": n.created_at.isoformat()
    } for n in qs]
    return JsonResponse({"ok": True, "notifications": notifs})


@require_http_methods(["POST"])
def api_notifications_mark_read(request):
    """Mark notification as read (student/teacher/admin)."""
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid json"}, status=400)

    nid = payload.get("id")
    if not nid:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    from users.models import Notification
    notif = get_object_or_404(Notification, id=nid, recipient=request.user)
    notif.is_read = True
    notif.save(update_fields=["is_read"])
    ActionLog.objects.create(user=request.user, action_type="Read notification", model_name="Notification", object_id=str(nid))
    return JsonResponse({"ok": True, "id": nid})



# old API -------------------------------------------------#


# -------------------------- No 2 Endpoint Ended -------------------------- #



# ---- USER MANAGEMENT ----
@login_required
@user_passes_test(is_admin)
def manage_users(request):
    users = User.objects.exclude(role="superadmin").order_by("-date_joined")
    return render(request, "exams/manage_users.html", {"users": users})


@login_required
@user_passes_test(is_admin)
def create_user(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        role = request.POST["role"]
        password = request.POST["password"]

        user = User.objects.create_user(username=username, email=email, password=password, role=role, approved=True)
        ActionLog.objects.create(user=request.user, action=f"Created user {user.username} ({role})")
        messages.success(request, f"User {username} created successfully.")
        return redirect("manage_users")

    return render(request, "exams/create_user.html")



# -------------------- CLASSES --------------------
def manage_classes_subjects(request):
    return render(request, 'exams/class_subject.html')

def serialize_all():
    """Return all classes and subjects in one dict"""
    classes = list(Class.objects.values("id", "name"))
    subjects = list(Subject.objects.values("id", "name", "school_class_id"))
    return {"classes": classes, "subjects": subjects}

@csrf_exempt
def class_subject_crud(request):
    if request.method == "GET":
        return JsonResponse(serialize_all(), safe=False)

    if request.method == "POST":
        data = json.loads(request.body)
        action = data.get("action")

        # ✅ CREATE CLASS
        if action == "create_class":
            Class.objects.create(name=data["name"])

        # ✅ UPDATE CLASS
        elif action == "update_class":
            try:
                cls = Class.objects.get(pk=data["id"])
                cls.name = data["name"]
                cls.save()
            except Class.DoesNotExist:
                return JsonResponse({"error": "Class not found"}, status=404)

        # ✅ DELETE CLASS
        elif action == "delete_class":
            Class.objects.filter(pk=data["id"]).delete()

        # ✅ CREATE SUBJECT
        elif action == "create_subject":
            Subject.objects.create(name=data["name"], school_class_id=data["class_id"])

        # ✅ UPDATE SUBJECT
        elif action == "update_subject":
            try:
                subj = Subject.objects.get(pk=data["id"])
                subj.name = data["name"]
                subj.school_class_id = data["class_id"]
                subj.save()
            except Subject.DoesNotExist:
                return JsonResponse({"error": "Subject not found"}, status=404)

        # ✅ DELETE SUBJECT
        elif action == "delete_subject":
            Subject.objects.filter(pk=data["id"]).delete()

        return JsonResponse(serialize_all(), safe=False)

    return JsonResponse({"error": "Invalid request"}, status=400)



# ---- QUIZ MANAGEMENT ----
@login_required
@user_passes_test(is_admin)
def manage_quizzes(request):
    quizzes = Quiz.objects.all().order_by("-created_at")

    page = Paginator(quizzes, 7)
    page_number = request.GET.get("page")
    page_obj = page.get_page(page_number)

    return render(request, "exams/manage_quizzes.html", {"page_obj": page_obj})



def search_quizzes(request):
    q = request.GET.get("q", "")
    page_number = request.GET.get("page", 1)

    quizzes = Quiz.objects.all()
    if q:
        quizzes = quizzes.filter(
            title__icontains=q
        ) | quizzes.filter(
            subject__name__icontains=q
        ) | quizzes.filter(
            subject__school_class__name__icontains=q
        )

    paginator = Paginator(quizzes.order_by("-start_time"), 5)  # 5 per page
    page_obj = paginator.get_page(page_number)

    data = {
        "results": [
            {
                "id": quiz.id,
                "title": quiz.title,
                "subject": quiz.subject.name,
                "class_name": quiz.subject.school_class.name,
                "start_time": quiz.start_time.strftime("%Y-%m-%d %H:%M"),
                "end_time": quiz.end_time.strftime("%Y-%m-%d %H:%M"),
                "is_published": quiz.is_published,
            }
            for quiz in page_obj
        ],
        "current_page": page_obj.number,
        "num_pages": paginator.num_pages,
        "has_previous": page_obj.has_previous(),
        "has_next": page_obj.has_next(),
        "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
        "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
    }
    return JsonResponse(data)



@login_required
@user_passes_test(is_admin)
def create_quiz(request):
    if request.method == "POST":
        title = request.POST["title"]
        subject_id = request.POST["subject"]
        duration = int(request.POST["duration"])

        subject = Subject.objects.get(id=subject_id)
        quiz = Quiz.objects.create(title=title, subject=subject, duration=duration, created_by=request.user)
        ActionLog.objects.create(user=request.user, action_type=f"Created quiz {quiz.title}")
        messages.success(request, f"Quiz {title} created successfully.")
        return redirect("manage_quizzes")

    subjects = Subject.objects.all()
    return render(request, "exams/create_quiz.html", {"subjects": subjects})


@login_required
@user_passes_test(is_admin)
def upload_quiz_excel(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        filepath = fs.path(filename)

        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
            title, subject_name, duration = row[:3]
            subject, _ = Subject.objects.get_or_create(name=subject_name)
            Quiz.objects.create(title=title, subject=subject, duration=int(duration), created_by=request.user)

        ActionLog.objects.create(user=request.user, action_type=f"Uploaded Exams from Excel")
        messages.success(request, "Exams uploaded successfully.")
        return redirect("manage_quizzes")

    return render(request, "exams/upload_quiz_excel.html")


# -----------------------------Student Quiz List with Status---------------------------------#  
from django.db.models import Q  

def get_quizzes_with_status(student):
    # Fetch all quizzes for student's class
    quizzes = Quiz.objects.filter(
        subject__school_class=student.student_class,
        created_by__role__in=["teacher", "admin", "superadmin"]
    ).select_related("subject", "created_by")

    # Map quiz attempts
    attempts = StudentQuizAttempt.objects.filter(student=student).select_related("quiz")
    attempts_map = {a.quiz_id: a for a in attempts}

    quizzes_with_status = []
    for quiz in quizzes:
        attempt = attempts_map.get(quiz.id)

        if attempt and attempt.is_submitted:
            if quiz.allow_retake:
                status = "Available for Retake"
            else:
                status = "Completed"
        else:
            status = "Not Started"

        quizzes_with_status.append({
            "quiz": quiz,
            "status": status
        })

    return quizzes_with_status



# -----------------------------Retake Approval & Request---------------------------------#


@login_required
@user_passes_test(is_admin)
def retake_requests_list(request):
    creator = User.objects.filter(role__in=['superadmin', 'admin', 'teacher']).first()
    print(creator)
    # Quizzes created by Superadmin, Admin, teacher
    quizzes = Quiz.objects.filter(created_by=creator).order_by("-start_time")
    print(quizzes)
    # Fetch attempts where retake is pending and student attempts on quizzes
    attempts = (
        StudentQuizAttempt.objects.filter(quiz__in=quizzes, retake_allowed=False, is_submitted=True)
        .select_related("student", "quiz")
        .order_by("-end_time")
    )

    retake_requests = (
        RetakeRequest.objects.filter(quiz__in=quizzes)
        .select_related("student", "quiz", "reviewed_by")
        .order_by("-created_at")
    )

    # Pagination
    paginator = Paginator(retake_requests, 10)  # 10 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "exams/admin_retake_requests.html", {"page_obj": page_obj, 'retake_requests': retake_requests, 'attempts': attempts})


@login_required
@user_passes_test(is_admin)
def approve_retake(request, quiz_id, student_id):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    quiz = get_object_or_404(Quiz, id=quiz_id)
    student = get_object_or_404(User, id=student_id, role="student")

    # find last attempt or create a fresh one
    attempt, created = StudentQuizAttempt.objects.get_or_create(
        student=student, quiz=quiz,
        defaults={"retake_allowed": True, "is_submitted": False}
    )

    # mark retake
    attempt.retake_allowed = True
    attempt.is_submitted = False  # reset if admin wants them to take again
    attempt.end_time = None
    attempt.retake_count += 1
    attempt.save()

    # log the action
    ActionLog.objects.create(
        user=request.user,
        action_type="Approved Retake",
        model_name="StudentQuizAttempt",
        object_id=str(attempt.id),
        details={"student": student.username, "exam": quiz.title, "retake_count": attempt.retake_count},
    )

    return JsonResponse({
        "success": True,
        "message": f"{student.username} can now retake {quiz.title}.",
        "retake_count": attempt.retake_count
    })



@login_required
@user_passes_test(is_admin_or_superadmin)
def handle_retake_request(request, request_id):
    retake_req = get_object_or_404(RetakeRequest, id=request_id, status="pending")
    decision = request.POST.get("decision")  # "approve" or "deny"

    if decision == "approve":
        # reset attempt or allow new
        attempt, created = StudentQuizAttempt.objects.get_or_create(
            student=retake_req.student,
            quiz=retake_req.quiz,
            defaults={"retake_allowed": True, "is_submitted": False, "retake_count": 1}
        )
        if not created:
            attempt.retake_allowed = True
            attempt.is_submitted = False
            attempt.end_time = None
            attempt.retake_count += 1
            attempt.save()

        retake_req.status = "approved"
        message = f"Your retake request for {retake_req.quiz.title} was approved."

    else:
        retake_req.status = "denied"
        message = f"Your retake request for {retake_req.quiz.title} was denied."

    retake_req.reviewed_by = request.user
    retake_req.reviewed_at = timezone.now()
    retake_req.save()

    # log
    ActionLog.objects.create(
        user=request.user,
        action_type=f"Retake {retake_req.status.capitalize()}",
        model_name="RetakeRequest",
        object_id=str(retake_req.id),
        details={"student": retake_req.student.username, "exam": retake_req.quiz.title}
    )

    # notify student
    Notification.objects.create(sender= request.user, recipient=retake_req.student, role='admin', message=message)

    return JsonResponse({"success": True, "message": message})

# -----------------------------Retake Approval & Request ended ---------------------------------#


@login_required
@user_passes_test(is_teacher_or_admin)
def teacher_dashboard(request):
    teacher = request.user

    # Quizzes created by this teacher
    quizzes = Quiz.objects.filter(created_by=teacher).order_by("-start_time")

    # Student attempts on teacher’s quizzes
    attempts = (
        StudentQuizAttempt.objects.filter(quiz__in=quizzes)
        .select_related("student", "quiz")
        .order_by("-end_time")
    )

    # Retake requests for teacher’s quizzes
    retake_requests = (
        RetakeRequest.objects.filter(quiz__in=quizzes)
        .select_related("student", "quiz", "reviewed_by")
        .order_by("-created_at")
    )

    return render(request, "exams/teacher_dashboard.html", {
        "quizzes": quizzes,
        "attempts": attempts,
        "retake_requests": retake_requests,
    })


# -------------------------Teacher Dashboard -------------------------------------#
# AJAX endpoints (JSON) for periodic refresh (every 10 seconds)
@login_required
def ajax_student_summary(request):
    user = request.user
    attempts = StudentQuizAttempt.objects.filter(student=user)
    total_attempts = attempts.count()
    avg_score = attempts.aggregate(avg=Avg('total_score'))['avg'] or 0
    pending_count = Answer.objects.filter(attempt__student=user, is_pending=True).count()
    data = {
        "total_attempts": total_attempts,
        "avg_score": float(avg_score),
        "pending_count": pending_count,
    }
    return JsonResponse(data)


@login_required
@user_passes_test(is_teacher)
def ajax_teacher_pending_subjectives(request):
    user = request.user
    pending_count = Answer.objects.filter(question__quiz__created_by=user, is_pending=True).count()
    return JsonResponse({"pending_count": pending_count})


# --------------------------------------------------------------#
# PDF export (consolidated results for a student)

def download_student_full_report(request, student_id):
    """Download full report for a single student (all quizzes)."""
    # Fetch attempts for this student
    attempts = StudentQuizAttempt.objects.filter(student_id=student_id).select_related("quiz", "quiz__subject")

    # Prepare PDF response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="student_{student_id}_results.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Header: logo | school name + student | student photo
    school_logo = os.path.join(settings.MEDIA_ROOT, "school_logo.png")
    student = attempts.first().student if attempts.exists() else None
    school_name = getattr(settings, "SCHOOL_NAME", "My School Name")
    student_photo = getattr(student, "profile_picture", None)  # assuming User has photo field

    header_data = [
        [
            Image(school_logo, width=50, height=50) if os.path.exists(school_logo) else "",
            Paragraph(f"<b>{school_name}</b><br/> " + (student.get_full_name() if student else "Unknown"), styles["Title"]),
            Image(student_photo.path, width=50, height=50) if student_photo and os.path.exists(student_photo.path) else "",
        ]
    ]
    header_table = Table(header_data, colWidths=[70, 350, 70])
    header_table.setStyle(TableStyle([("ALIGN", (1, 0), (1, 0), "CENTER")]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))

    # Loop through attempts
    for attempt in attempts:
        elements.append(Paragraph(f"<b>Exam:</b> {attempt.quiz.title}", styles["Heading3"]))
        elements.append(Paragraph(f"Subject: {attempt.quiz.subject.name}", styles["Normal"]))
        elements.append(Paragraph(f"Date: {attempt.started_at.strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))

        # Collect answers
        answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice")

        data = [["Question", "Answer", "Marks", "Feedback"]]
        for ans in answers:
            if ans.selected_choice:
                ans_text = ans.selected_choice.text
            else:
                ans_text = ans.text_answer or "-"
            data.append([ans.question.text, ans_text, ans.obtained_marks, ans.feedback or ""])

        table = Table(data, colWidths=[200, 150, 60, 100])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 15))

    doc.build(elements)
    return response


def download_closed_quiz_report(request, quiz_id):
    """Download report for all students who attempted a closed quiz."""
    quiz = Quiz.objects.get(id=quiz_id)
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related("student")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="quiz_{quiz_id}_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Header
    school_logo = os.path.join(settings.MEDIA_ROOT, "school_logo.png")
    school_name = getattr(settings, "SCHOOL_NAME", "My school name")
    header_data = [
        [
            Image(school_logo, width=50, height=50) if os.path.exists(school_logo) else "",
            Paragraph(f"<b>{school_name}</b><br/>Exam Report: {quiz.title}", styles["Title"]),
            "",
        ]
    ]
    header_table = Table(header_data, colWidths=[70, 350, 70])
    header_table.setStyle(TableStyle([("ALIGN", (1, 0), (1, 0), "CENTER")]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))

    # Loop through each student's attempt
    for attempt in attempts:
        student = attempt.student
        elements.append(Paragraph(f"<b>Student:</b> {student.get_full_name()} ({getattr(student, 'student_class', 'N/A')})", styles["Heading3"]))
        elements.append(Paragraph(f"Date: {attempt.started_at.strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))

        answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice")

        data = [["Question", "Answer", "Marks", "Feedback"]]
        for ans in answers:
            if ans.selected_choice:
                ans_text = ans.selected_choice.text
            else:
                ans_text = ans.text_answer or "-"
            data.append([ans.question.text, ans_text, ans.obtained_marks, ans.feedback or ""])

        table = Table(data, colWidths=[200, 150, 60, 100])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 15))

    doc.build(elements)
    return response


# -------------------------
# Leaderboard (global top students by average score)
# -------------------------
@login_required
def leaderboard(request):
    """
    Renders a leaderboard template showing top students by average score across their completed attempts.
    """
    # compute average attempt scores per student; use StudentQuizAttempt.score (we ensure it's set on submit)
    qs = StudentQuizAttempt.objects.filter(is_submitted=True).values("student").annotate(
        avg_score=Avg("score"), attempts=Sum("score")
    ).order_by("-avg_score")[:50]

    # map to user objects + scores
    leaderboard_list = []
    for item in qs:
        user_id = item["student"]
        try:
            user = User.objects.get(id=user_id)
            leaderboard_list.append({"student": user, "avg_score": item["avg_score"] or 0})
        except User.DoesNotExist:
            continue

    return render(request, "exams/leaderboard.html", {"leaderboard": leaderboard_list})



# ---- Page: create quiz (manual + excel upload) ----

def is_teacher_or_admin(user):
    return user.is_authenticated and user.role in ('teacher','admin','superadmin')


@login_required
@user_passes_test(is_teacher_or_admin)
def create_quiz_page(request):
    subjects = Subject.objects.select_related('school_class').all()
    return render(request, "exams/create_quiz.html", {"subjects": subjects})

# ---- AJAX JSON create ----
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def create_quiz_ajax(request):
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    title = payload.get("title")
    subject_id = payload.get("subject_id")
    start_time = payload.get("start_time")
    end_time = payload.get("end_time")
    duration_minutes = payload.get("duration_minutes") or 30
    is_published = bool(payload.get("is_published", False))
    questions = payload.get("questions", [])

    if not title or not subject_id:
        return JsonResponse({"ok": False, "error": "Title and subject are required."}, status=400)
    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Subject not found."}, status=404)

    # parse datetimes (accept ISO or 'YYYY-MM-DDTHH:MM' from datetime-local)
    def parse_dt(s):
        if not s:
            return None
        try:
            if "T" in s:
                s = s.replace("T", " ")
            return timezone.make_aware(datetime.strptime(s, "%Y-%m-%d %H:%M"))
        except Exception:
            try:
                return timezone.make_aware(datetime.fromisoformat(s))
            except Exception:
                return None

    st = parse_dt(start_time)
    et = parse_dt(end_time)
    if not st or not et:
        return JsonResponse({"ok": False, "error": "Invalid start_time or end_time format. Use YYYY-MM-DD HH:MM"}, status=400)

    # Build quiz within transaction
    try:
        with transaction.atomic():
            quiz = Quiz.objects.create(
                title=title,
                subject=subject,
                created_by=request.user,
                start_time=st,
                end_time=et,
                duration_minutes=int(duration_minutes),
                is_published=is_published
            )
            for qidx, q in enumerate(questions):
                qtext = q.get("text")
                qtype = q.get("question_type")
                qmarks = q.get("marks", 1)
                if not qtext or qtype not in ("objective","subjective"):
                    raise ValueError(f"Invalid question at index {qidx}")

                question = Question.objects.create(
                    quiz=quiz,
                    text=qtext,
                    question_type=qtype,
                    marks=int(qmarks)
                )

                if qtype == "objective":
                    choices = q.get("choices", [])
                    if not choices or not isinstance(choices, list):
                        raise ValueError(f"Objective question requires choices at index {qidx}")
                    correct_present = False
                    for cidx, c in enumerate(choices):
                        ctext = c.get("text")
                        cis = bool(c.get("is_correct", False))
                        if not ctext:
                            raise ValueError(f"Choice text missing for question {qidx} choice {cidx}")
                        Choice.objects.create(question=question, text=ctext, is_correct=cis)
                        if cis:
                            correct_present = True
                    if not correct_present:
                        raise ValueError(f"At least one correct choice required for question {qidx}")
            # success
            return JsonResponse({"ok": True, "quiz_id": quiz.id, "message": "Quiz created."})
    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Server error: " + str(e)}, status=500)

# ---- Excel import ----
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def import_quiz_excel(request):
    """
    Expect FormData with 'excel_file' and optionally 'subject_id' (or subject/class declared in file).
    Excel format (first sheet):
      - Row 1..6: metadata key/value in column A/B:
         A1: quiz_title  B1: My Quiz Title
         A2: class_name  B2: JSS1
         A3: subject_name B3: Mathematics
         A4: start_time  B4: 2025-09-13 14:00
         A5: end_time    B5: 2025-09-13 15:00
         A6: duration_minutes B6: 60
         A7: is_published B7: True
      - Row 9 header then rows from 10:
         Columns: question_text | question_type | marks | choice_1 | choice_1_correct (0/1) | choice_2 | choice_2_correct | ...
    """
    f = request.FILES.get('excel_file')
    if not f:
        return JsonResponse({"ok": False, "error": "No file uploaded"}, status=400)

    # save to temp location (openpyxl can read file-like objects but safer to store)
    try:
        wb = load_workbook(filename=f, data_only=True)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Invalid Excel file: " + str(e)}, status=400)

    sheet = wb.active

    def cell_value(r,c):
        return sheet.cell(row=r, column=c).value

    # read metadata
    try:
        meta = {}
        for r in range(1, 9):
            key = cell_value(r,1)
            val = cell_value(r,2)
            if key:
                meta[str(key).strip().lower()] = val
    except Exception:
        return JsonResponse({"ok": False, "error": "Failed to read metadata"}, status=400)

    title = meta.get("exam_title") or meta.get("title")
    class_name = meta.get("class_name")
    subject_name = meta.get("subject_name")
    start_time_s = meta.get("start_time")
    end_time_s = meta.get("end_time")
    duration = int(meta.get("duration_minutes") or meta.get("duration") or 30)
    published = bool(meta.get("is_published")) if meta.get("is_published") is not None else False

    # validate subject/class
    if not (class_name and subject_name and title):
        return JsonResponse({"ok": False, "error": "Metadata must include class_name, subject_name and exam_title"}, status=400)

    try:
        school_class = Class.objects.get(name__iexact=str(class_name).strip())
    except Class.DoesNotExist:
        return JsonResponse({"ok": False, "error": f"Class '{class_name}' not found"}, status=404)
    try:
        subject = Subject.objects.get(name__iexact=str(subject_name).strip(), school_class=school_class)
    except Subject.DoesNotExist:
        return JsonResponse({"ok": False, "error": f"Subject '{subject_name}' not found for class {class_name}"}, status=404)

    # parse datetimes (support datetime objects too)
    def parse_dt_obj(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return timezone.make_aware(v) if timezone.is_naive(v) else v
        try:
            s = str(v)
            if "T" in s: s = s.replace("T"," ")
            return timezone.make_aware(datetime.strptime(s, "%Y-%m-%d %H:%M"))
        except Exception:
            try:
                return timezone.make_aware(datetime.fromisoformat(str(v)))
            except Exception:
                return None

    st = parse_dt_obj(start_time_s)
    et = parse_dt_obj(end_time_s)
    if not st or not et:
        return JsonResponse({"ok": False, "error": "Invalid start_time or end_time in metadata. Use 'YYYY-MM-DD HH:MM' or Excel datetime."}, status=400)

    # find header row (we expect header at row 9 or row with 'question_text')
    header_row = None
    for r in range(1, 30):
        first_col = cell_value(r,1)
        if first_col and str(first_col).strip().lower() in ("question_text","question","q_text"):
            header_row = r
            break
    if header_row is None:
        # assume header at row 9
        header_row = 9

    # parse questions starting from header_row+1 to last row with question_text
    questions = []
    r = header_row + 1
    while True:
        q_text = cell_value(r, 1)
        if q_text is None:
            break
        q_type = cell_value(r, 2) or 'objective'
        q_marks = int(cell_value(r, 3) or 1)

        # choices start at col 4, every pair: text, correct
        choices = []
        col = 4
        while True:
            c_text = cell_value(r, col)
            c_flag = cell_value(r, col+1)
            if c_text is None:
                break
            is_corr = False
            if c_flag in (1, '1', True, 'TRUE', 'true'):
                is_corr = True
            choices.append({"text": str(c_text).strip(), "is_correct": bool(is_corr)})
            col += 2

        questions.append({
            "text": str(q_text).strip(),
            "question_type": str(q_type).strip().lower(),
            "marks": q_marks,
            "choices": choices
        })
        r += 1

    # Now create quiz
    try:
        with transaction.atomic():
            quiz = Quiz.objects.create(
                title=title,
                school_class=school_class,
                subject=subject,
                created_by=request.user,
                start_time=st,
                end_time=et,
                duration_minutes=duration,
                is_published=published
            )
            # create questions
            for q in questions:
                qtext = q["text"]
                qtype = q["question_type"]
                qmarks = q["marks"]
                question = Question.objects.create(quiz=quiz, text=qtext, question_type=qtype, marks=qmarks)
                if qtype == "objective":
                    if not q["choices"]:
                        raise ValueError("Objective question without choices found in Excel.")
                    correct_found = False
                    for c in q["choices"]:
                        Choice.objects.create(question=question, text=c["text"], is_correct=c["is_correct"])
                        if c["is_correct"]:
                            correct_found = True
                    if not correct_found:
                        raise ValueError("Objective question must have at least one correct choice.")
            return JsonResponse({"ok": True, "quiz_id": quiz.id, "message": "Exam imported from Excel."})
    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Import failed: " + str(e)}, status=500)


# ---- Manage quizzes page ----
@login_required
@user_passes_test(is_teacher_or_admin)
def manage_quizzes_page(request):
    if request.user.role == 'teacher':
        quizzes = Quiz.objects.filter(created_by=request.user).order_by('-created_at')
    else:
        quizzes = Quiz.objects.all().order_by('-created_at')
    return render(request, "exams/manage_quizzes.html", {"quizzes": quizzes})

# ---- AJAX publish toggle ---old first----#
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def publish_toggle_ajax(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # only creator or admin can toggle
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)
    quiz.is_published = not quiz.is_published
    quiz.save()
    return JsonResponse({"ok": True, "is_published": quiz.is_published})

# ---- AJAX delete ----
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def delete_quiz_ajax(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)
    quiz.delete()
    ActionLog.objects.create(
    user=request.user,
    action_type= "Delete Exam",
    model_name="Exam",
    object_id=str(quiz.id),
    details={"title": quiz.title, "subject": quiz.subject.name, "action": "deleted"},
)

    return JsonResponse({"ok": True})



def is_teacher_or_admin(user):
    return user.is_authenticated and user.role in ('teacher','admin','superadmin')

@login_required
@user_passes_test(is_teacher_or_admin)
def edit_quiz_page(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # permission: only creator or admin allowed to edit
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return HttpResponse("Permission denied", status=403)
    subjects = Subject.objects.select_related('school_class').all()

    # We'll render a page similar to create_quiz but include a small script that fetches quiz JSON to prefill
    return render(request, "exams/edit_quiz.html", {"quiz": quiz, "subjects": subjects})



@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def edit_quiz_ajax(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    title = payload.get("title")
    subject_id = payload.get("subject_id")
    start_time = payload.get("start_time")
    end_time = payload.get("end_time")
    duration_minutes = payload.get("duration_minutes") or 30
    is_published = bool(payload.get("is_published", False))
    questions = payload.get("questions", [])

    if not title or not subject_id:
        return JsonResponse({"ok": False, "error": "Title and subject are required."}, status=400)

    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Subject not found."}, status=404)

    # Update within transaction. We'll replace existing questions/choices with new ones.
    try:
        with transaction.atomic():
            quiz.title = title
            quiz.subject = subject
            # parse start/end strings same as in create; expecting ISO or "YYYY-MM-DD HH:MM"
            from django.utils.dateparse import parse_datetime

            start_time = parse_datetime(payload.get("start_time"))
            end_time = parse_datetime(payload.get("end_time"))

            if not start_time or not end_time:
                return JsonResponse({"ok": False, "error": "Invalid start/end time format."}, status=400) 
            
            quiz.start_time = start_time
            quiz.end_time = end_time
            quiz.duration_minutes = int(duration_minutes)
            quiz.is_published = is_published
            quiz.save()
            

            # delete old questions & choices, then recreate
            quiz.questions.all().delete()
            for qidx, q in enumerate(questions):
                qtext = q.get("text")
                qtype = q.get("question_type")
                qmarks = q.get("marks", 1)
                if not qtext or qtype not in ("objective","subjective"):
                    raise ValueError(f"Invalid question at index {qidx}")
                question = Question.objects.create(
                    quiz=quiz, text=qtext, question_type=qtype, marks=int(qmarks)
                )
                if qtype == "objective":
                    choices = q.get("choices", [])
                    if not choices or not isinstance(choices, list):
                        raise ValueError(f"Objective question requires choices at index {qidx}")
                    correct_found = False
                    for cidx, c in enumerate(choices):
                        ctext = c.get("text")
                        cis = bool(c.get("is_correct", False))
                        if not ctext:
                            raise ValueError(f"Choice text missing for question {qidx} choice {cidx}")
                        Choice.objects.create(question=question, text=ctext, is_correct=cis)
                        if cis:
                            correct_found = True
                    if not correct_found:
                        raise ValueError(f"At least one correct choice required for question {qidx}")
            # success
            ActionLog.objects.create(
            user=request.user,
            action_type="Edit Exam",
            model_name="Exam",
            object_id=str(quiz.id),
            details={"title": quiz.title, "subject": quiz.subject.name, "action": "edited" },
        )

        return JsonResponse({"ok": True, "quiz_id": quiz.id, "message": "Quiz updated."})
    
    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Server error: " + str(e)}, status=500)



@require_POST
def toggle_quiz_publish(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    quiz.is_published = not quiz.is_published
    quiz.save()
    return JsonResponse({"ok": True, "new_status": quiz.is_published})

def quiz_detail_api(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    return JsonResponse({
        "ok": True,
        "quiz": {
            "id": quiz.id,
            "title": quiz.title,
            "subject": str(quiz.subject),
            "class_name": str(quiz.subject.school_class),
            "duration_minutes": quiz.duration_minutes,
            "is_published": quiz.is_published,
        }
    })


@login_required
@user_passes_test(is_teacher_or_admin)
def quiz_details_page(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # permission: only creator or admin allowed to view details
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return HttpResponse("Permission denied", status=403)
    if quiz.DoesNotExist:
        attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related('student').order_by('-started_at')
        return redirect('quiz_closed_detail', quiz.id)
    
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related('student').order_by('-started_at')

    return render(request, "exams/quiz_details.html", {"quiz": quiz, "attempts": attempts})




def quiz_closed(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    # Make sure quiz is actually closed
    if quiz.end_time > timezone.now():
        # If not yet closed, you can redirect to quiz detail or show error
        return render(request, "exams/not_closed.html", {"quiz": quiz})
    
    # Get all attempts
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related("student").order_by("-started_at")
    context = {
        "quiz": quiz,
        "attempts": attempts,
    }
    return render(request, "exams/quiz_closed.html", context)




from openpyxl import Workbook
from django.http import HttpResponse

@login_required
@user_passes_test(is_teacher_or_admin)
def download_excel_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "ExamTemplate"

    # metadata rows
    ws['A1'] = 'exam_title'; ws['B1'] = settings.SCHOOL_NAME + " " + "FIRST TERM EXAM"
    ws['A2'] = 'class_name'; ws['B2'] = 'JSS1'
    ws['A3'] = 'subject_name'; ws['B3'] = 'Mathematics'
    ws['A4'] = 'start_time'; ws['B4'] = None
    ws['A5'] = 'end_time'; ws['B5'] = None
    ws['A6'] = 'duration_minutes'; ws['B6'] = 60
    ws['A7'] = 'is_published'; ws['B7'] = 'True'

    # header row at row 9
    headers = ['question_text','question_type','marks']
    # allow up to 6 choices (choice_x, choice_x_correct)
    for i in range(1,7):
        headers.append(f'choice_{i}')
        headers.append(f'choice_{i}_correct')
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=9, column=col_idx, value=header)

    # example question row
    ws.cell(row=10, column=1, value='What is 2+2?')
    ws.cell(row=10, column=2, value='objective')
    ws.cell(row=10, column=3, value='1')
    ws.cell(row=10, column=4, value='3')
    ws.cell(row=10, column=5, value='0')
    ws.cell(row=10, column=6, value='4')
    ws.cell(row=10, column=7, value='1')

    # prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Exam_template.xlsx'
    wb.save(response)
    return response


@login_required
def quiz_json3(request, attempt_id):
    """
    Return quiz questions + answers as JSON for a student's attempt.
    Supports resume if attempt not submitted.
    """

    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)

    # Check expiry
    if attempt.end_time and timezone.now() > attempt.end_time:
        return JsonResponse({"error": "Exam time expired."}, status=403)

    # Check submission
    if attempt.is_submitted:
        return JsonResponse({"error": "Exam already submitted."}, status=403)
    quiz = attempt.quiz
    questions = quiz.questions.prefetch_related("choice_set")

    data = []
    for q in questions:
        saved_answer = Answer.objects.filter(attempt=attempt, question=q).first()
        data.append({
            "id": q.id,
            "text": q.text,
            "type": q.question_type,
            "marks": q.marks,
            "choices": [{"id": c.id, "text": c.text} for c in q.choice_set.all()],
            "saved_choice": saved_answer.selected_choice.id if saved_answer and saved_answer.selected_choice else None,
            "saved_text": saved_answer.text_answer if saved_answer else "",
        })


    return JsonResponse({
        "quiz": {
            "id": quiz.id,
            "title": quiz.title,
            "subject": quiz.subject.name,
            "duration_minutes": quiz.duration,
            "end_time": attempt.end_time.isoformat() if attempt.end_time else None,
        },
        "questions": data,
    })




@login_required
def quiz_json1(request, quiz_id):
    quiz = get_object_or_404(Quiz.objects.prefetch_related("choices"), id=quiz_id)

    data = {
        "quiz": {
            "id": quiz.id,
            "title": quiz.title,
            "end_time": quiz.end_time.isoformat() if quiz.end_time else None,
            "questions": []
        }
    }

    for q in quiz.questions.all():
        data["quiz"]["questions"].append({
            "id": q.id,
            "text": q.text,
            "question_type": q.question_type,
            "marks": q.marks,
            "choices": [
                {"id": c.id, "text": c.text} for c in q.choices.all()
            ]
        })

    return JsonResponse(data)


@login_required

def quiz_json_quiz_load(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)

    # permission: only creator or admin/superadmin can fetch edit JSON
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

    quiz_data = {
        "id": quiz.id,
        "title": quiz.title,
        "subject_id": quiz.subject.id if quiz.subject else None,
        "subject_name": quiz.subject.name if quiz.subject else None,
        "start_time": quiz.start_time.strftime("%Y-%m-%d %H:%M") if quiz.start_time else None,
        "end_time": quiz.end_time.strftime("%Y-%m-%d %H:%M") if quiz.end_time else None,
        "duration_minutes": quiz.duration_minutes,
        "is_published": quiz.is_published,
        "questions": []
    }
    for q in quiz.questions.all():
        qdata = {
            "id": q.id,
            "text": q.text,
            "question_type": q.question_type,
            "marks": q.marks,
            "choices": []  
        }
        if q.question_type == "objective":
            qdata["choices"] = [
                {"id": c.id, "text": c.text, "is_correct": c.is_correct}
                for c in q.choices.all()
            ]
        quiz_data["questions"].append(qdata)
      

    return JsonResponse({"ok": True, "quiz": quiz_data})


###------- Retake Exam after submission by student this will check their attempt and retake count-------###

@login_required
def request_retake(request, attempt_id):
    """
    Student requests a retake for a submitted attempt.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)

    if not attempt.completed:
        messages.error(request, "You can only request a retake after submitting the Exam.")
        return redirect("student_dashboard")

    if getattr(attempt, "retake_requested", False):
        messages.info(request, "You have already requested a retake for this Exam.")
        return redirect("student_dashboard")

    # Mark request
    attempt.retake_requested = True
    attempt.save()

    # Notify admin/teacher
    Notification.objects.create(
        sender=request.user,
        recipient=attempt.quiz.created_by,
        message=f"📩 {request.user.username} has requested a retake for '{attempt.quiz.title}'.",
        role="admin",
        is_broadcast=False,
    )

    messages.success(request, "Your retake request has been sent for approval.")
    return redirect("student_dashboard")

###---------------------The End Retake Exam after submission -----------------------###



###------- Retake Exam after submission by student this will check their attempt and retake count-------###

@user_passes_test(is_admin_or_superadmin)
def retake_requests(request): # Retake request button on dashboard
    """
    Admin page: list attempts that are submitted/expired.
    """
    attempts = StudentQuizAttempt.objects.filter(is_submitted=True).select_related("quiz", "student")
    return render(request, "exams/retake_requests.html", {"attempts": attempts})


@user_passes_test(is_admin_or_superadmin)
def approve_retake(request, attempt_id):
    """
    Admin approves retake request.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)

    # Reset attempt for retake
    attempt.retake_allowed = True
    attempt.is_submitted = False
    attempt.end_time = None
    attempt.save()

    # Notify studenl
    Notification.objects.create(
        recipient=attempt.student,
        message=f"✅ Your retake request for Exam '{attempt.quiz.title}' has been approved.",
        is_read=False
    )

    # Log action
    ActionLog.objects.create(
        user=request.user,
        action_type="approve_retake",
        details={"attempt_id": attempt.id, "student": attempt.student.username, "Exam": attempt.quiz.title}
    )

    messages.success(request, f"Retake approved for {attempt.student.username} on {attempt.quiz.title}.")
    return redirect("retake_requests")



##------------New Student Exam take/Retake/ etc Page ------------------###



# helper: check student
def is_student(user):
    return user.is_authenticated and getattr(user, "role", None) == "student"

def is_teacher_or_admin(user):
    return user.is_authenticated and getattr(user, "role", None) in ("teacher", "admin", "superadmin")


@login_required
def take_quiz_view(request, quiz_id):
    """
    Show the take-quiz page. Create or resume attempt server-side.
    """
    if not is_student(request.user):
        return HttpResponseForbidden("forbidden")

    quiz = get_object_or_404(Quiz, id=quiz_id)

    now = timezone.now()
    # allowed to start/resume if published OR retake allowed by admin/teacher, etc.
    # find existing active attempt
    attempt = StudentQuizAttempt.objects.filter(student=request.user, quiz=quiz, is_submitted=False).order_by("-started_at").first()

    if attempt:
        # if expired, do not resume (unless retake_allowed)
        if attempt.end_time and now > attempt.end_time and not attempt.retake_allowed:
            attempt = None

    if not attempt:
        # not resuming: check whether quiz open or retake allowed
        available = quiz.is_published and (quiz.start_time <= now <= quiz.end_time)
        last_attempt = StudentQuizAttempt.objects.filter(student=request.user, quiz=quiz).order_by("-started_at").first()
        allow_ret = quiz.allow_retake or (last_attempt and last_attempt.retake_allowed)
        if not available and not allow_ret:
            # if closed and not allowed: if there is a is_submitted attempt, redirect to its result
            if last_attempt and last_attempt.is_submitted:
                return redirect('quiz_result', attempt_id=last_attempt.id)
            return redirect('quiz_closed_detail', quiz_id=quiz.id)

        # create new attempt
        end_time = now + timezone.timedelta(minutes=quiz.duration_minutes)
        attempt = StudentQuizAttempt.objects.create(
            student=request.user,
            quiz=quiz,
            started_at=now,
            end_time=end_time,
            is_submitted=False,
            retake_allowed=False,
            retake_count=(last_attempt.retake_count+1 if last_attempt else 0),
            score=0.0
        )
        ActionLog.objects.create(user=request.user, action_type="Started Exam", description=f"Started exam {quiz.title}", model_name="StudentQuizAttempt", object_id=str(attempt.id), details={"quiz": quiz.title})

    # ensure questions are prefetched for template rendering
    questions = quiz.questions.prefetch_related("choices").all()
    print(questions)
    # also pass existing saved answers to prefill (dict question_id -> answer)
    saved_answers = {}
    for ans in Answer.objects.filter(attempt=attempt).select_related("selected_choice", "question"):
        if ans.question.question_type == "objective" and ans.selected_choice:
            saved_answers[str(ans.question.id)] = {"choice_id": ans.selected_choice.id}
        else:
            saved_answers[str(ans.question.id)] = {"text": ans.text_answer}

    context = {
        "quiz": quiz,
        "attempt": attempt,
        "questions": questions,
        "saved_answers": saved_answers,
    }
    return render(request, "exams/take_quiz.html", context)


@login_required
@require_POST
def api_submit_answer(request, attempt_id):
    """
    Autosave single answer (AJAX). Accepts JSON: {question_id, answer}
    - For objective: answer is choice id (int)
    - For subjective: answer is text
    Autograde objectives here.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)
    if attempt.is_submitted:
        return JsonResponse({"ok": False, "error": "Attempt already submitted"}, status=400)
    if attempt.end_time and timezone.now() > attempt.end_time:
        return JsonResponse({"ok": False, "error": "Attempt time expired"}, status=400)

    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid json"}, status=400)

    qid = payload.get("question_id")
    answer = payload.get("answer", None)  # can be choice id or text

    if not qid:
        return JsonResponse({"ok": False, "error": "question_id required"}, status=400)

    question = get_object_or_404(Question, id=qid, quiz=attempt.quiz)

    # objective
    if question.question_type == "objective":
        # answer should be a choice id (int or str)
        try:
            choice = Choice.objects.get(id=int(answer), question=question)
        except Exception:
            choice = None

        if choice:
            obtained = float(question.marks) if choice.is_correct else 0.0
            ans_obj, created = Answer.objects.update_or_create(
                attempt=attempt, question=question,
                defaults={
                    "selected_choice": choice,
                    "text_answer": None,
                    "obtained_marks": obtained,
                    "is_pending": False,
                }
            )
        else:
            # clear selection
            ans_obj, created = Answer.objects.update_or_create(
                attempt=attempt, question=question,
                defaults={
                    "selected_choice": None,
                    "text_answer": None,
                    "obtained_marks": 0.0,
                    "is_pending": False,
                }
            )
    else:
        # subjective: save text, mark is_pending True, obtained_marks left as 0 (teacher will grade later)
        text = str(answer or "")
        ans_obj, created = Answer.objects.update_or_create(
            attempt=attempt, question=question,
            defaults={
                "selected_choice": None,
                "text_answer": text,
                "obtained_marks": 0.0,
                "is_pending": True,
            }
        )

    ActionLog.objects.create(user=request.user, action_type="submit_answer", description="Autosaved answer", model_name="Answer", object_id=str(ans_obj.id), details={"question": question.id})
    return JsonResponse({"ok": True})


@login_required
@require_POST
def api_submit_attempt(request, attempt_id):
    """
    Final submission: autograde objective parts (again to be safe), sum objective + graded subjective.
    Returns JSON with overall score and per-question details.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)

    if attempt.is_submitted:
        return JsonResponse({"ok": False, "error": "Already submitted"}, status=400)

    # re-evaluate objective answers and ensure answers exist for all objective questions
    answers_qs = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice")
    # Ensure every objective question has an Answer row (create with 0 if missing)
    for q in attempt.quiz.questions.filter(question_type="objective"):
        ans, created = Answer.objects.get_or_create(attempt=attempt, question=q, defaults={
            "selected_choice": None, "text_answer": None, "obtained_marks": 0.0, "is_pending": False
        })

    # Auto-grade objective answers
    total_score = 0.0
    for ans in Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice"):
        q = ans.question
        if q.question_type == "objective":
            if ans.selected_choice and getattr(ans.selected_choice, "is_correct", False):
                ans.obtained_marks = float(q.marks)
            else:
                ans.obtained_marks = 0.0
            ans.is_pending = False
            ans.save(update_fields=["obtained_marks", "is_pending"])
        # subjective answers left as is (pending until graded)

    # calculate totals using Answer helper methods
    objective_sum = Answer.objective_score(attempt)
    subjective_sum = Answer.subjective_score(attempt)  # only graded subjective answers contribute
    total_score = float(objective_sum) + float(subjective_sum)

    attempt.score = total_score
    attempt.is_submitted = True
    attempt.submitted_at = timezone.now() if hasattr(attempt, "submitted_at") else timezone.now()
    attempt.save(update_fields=["score", "is_submitted", "submitted_at"])

    ActionLog.objects.create(user=request.user, action_type="Submitted Exam", description=f"Submitted attempt {attempt.id}", model_name="StudentQuizAttempt", object_id=str(attempt.id), details={"score": attempt.score})

    # Build a summary payload
    question_details = []
    for ans in Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice"):
        q = ans.question
        if q.question_type == "objective":
            student_ans = ans.selected_choice.text if ans.selected_choice else None
            correct = bool(ans.selected_choice and getattr(ans.selected_choice, "is_correct", False))
            correct_ans_text = ", ".join([c.text for c in q.choices.filter(is_correct=True)])
            question_details.append({
                "question_id": q.id,
                "type": "objective",
                "text": q.text,
                "student": student_ans,
                "correct_answer": correct_ans_text,
                "marks_awarded": ans.obtained_marks,
                "max_marks": q.marks,
                "status": "correct" if correct else "wrong",
            })
        else:
            question_details.append({
                "question_id": q.id,
                "type": "subjective",
                "text": q.text,
                "student": ans.text_answer,
                "marks_awarded": ans.obtained_marks if not ans.is_pending else None,
                "max_marks": q.marks,
                "status": "pending" if ans.is_pending else "graded",
                "feedback": ans.feedback,
            })

    return JsonResponse({
        "ok": True,
        "score": total_score,
        "objective_sum": float(objective_sum),
        "subjective_sum": float(subjective_sum),
        "questions": question_details,
    })


@login_required
def quiz_result_view(request, attempt_id):
    """
    Render result page with per-question breakdown and Chart.js bar for objective correctness.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)
    # allow only owner or teacher/admin
    if request.user != attempt.student and not is_teacher_or_admin(request.user):
        return HttpResponseForbidden("forbidden")

    answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice", "graded_by")
    qrows = []
    correct_count = 0
    for ans in answers:
        q = ans.question
        if q.question_type == "objective":
            correct = bool(ans.selected_choice and getattr(ans.selected_choice, "is_correct", False))
            if correct:
                correct_count += 1
            correct_text = ", ".join([c.text for c in q.choices.filter(is_correct=True)])
            student_text = ans.selected_choice.text if ans.selected_choice else (ans.text_answer or "")
            qrows.append({
                "id": q.id,
                "text": q.text,
                "type": "objective",
                "student_answer": student_text,
                "correct_answer": correct_text,
                "marks_awarded": ans.obtained_marks,
                "max_marks": q.marks,
                "status": "correct" if correct else "wrong",
            })
        else:
            qrows.append({
                "id": q.id,
                "text": q.text,
                "type": "subjective",
                "student_answer": ans.text_answer,
                "marks_awarded": (ans.obtained_marks if not ans.is_pending else None),
                "max_marks": q.marks,
                "status": "pending" if ans.is_pending else "graded",
                "feedback": ans.feedback,
            })

    # chart data for objective questions: percent correct per question (0 or 1)
    chart_labels = []
    chart_values = []
    for row in qrows:
        chart_labels.append(f"Q{row['id']}")
        if row['type'] == 'objective':
            chart_values.append((row['marks_awarded'] or 0) / (row['max_marks'] or 1) * 100)
        else:
            chart_values.append(None)  # will show as gap

    context = {
        "attempt": attempt,
        "qrows": qrows,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "total_score": attempt.score,
        "correct_count": correct_count,
        "total_questions": len(qrows),
    }
    return render(request, "exams/quiz_result.html", context)


@login_required
def review_attempt_view(request, attempt_id):
    """
    Review answers + teacher feedback. Owner or teacher/admin only.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)
    if request.user != attempt.student and not is_teacher_or_admin(request.user):
        return HttpResponseForbidden("forbidden")

    answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice", "graded_by")
    return render(request, "exams/quiz_review.html", {"attempt": attempt, "answers": answers})



@login_required
@require_POST
def request_retake_view(request, quiz_id):
    """
    Student requests retake: creates RetakeRequest and notifies teacher + admins.
    """
    if not is_student(request.user):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    quiz = get_object_or_404(Quiz, id=quiz_id)

    # Get reason (from JSON body or POST)
    reason = ""
    try:
        payload = json.loads(request.body.decode())
        reason = payload.get("reason", "")[:1000]
    except Exception:
        reason = request.POST.get("reason", "")[:1000]

    # Create RetakeRequest
    rr = RetakeRequest.objects.create(
        student=request.user,
        quiz=quiz,
        reason=reason,
        status="pending"
    )

    # Notify teacher
    Notification.objects.create(
        sender=request.user,
        recipient=quiz.created_by,
        message=f"Retake request by {request.user.get_full_name()} for {quiz.title}: {reason}",
        role="teacher",
        is_broadcast=False,
    )

    # Notify admins
    admins = User.objects.filter(role__in=("admin", "superadmin"), approved=True)
    for a in admins:
        Notification.objects.create(
            sender=request.user,
            recipient=a,
            message=f"Retake request by {request.user.get_full_name()} for {quiz.title}",
            role="admin",
            is_broadcast=False,
        )

    # Log action
    ActionLog.objects.create(
        user=request.user,
        action_type="retake_request",
        description="Requested retake",
        model_name="RetakeRequest",
        object_id=str(rr.id),
        details={"quiz": quiz.title},
    )

    return JsonResponse({"ok": True, "message": "Retake request sent."})


@login_required
@require_POST
def approve_request_retake_view(request, req_id):
    """
    Teacher/admin approves a retake request. This sets the last attempt.retake_allowed True (or creates one)
    and notifies the student.
    """
    if not is_teacher_or_admin(request.user):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    rr = get_object_or_404(RetakeRequest, id=req_id)
    if rr.status == "approved":
        return JsonResponse({"ok": False, "error": "Already approved"}, status=400)

    rr.status = "approved"
    rr.reviewed_at = timezone.now()
    rr.reviewed_by = request.user
    rr.save(update_fields=["status", "reviewed_at", "reviewed_by"])

    # grant retake: find last attempt and mark retake_allowed True. If none, create a new attempt with retake_allowed True
    last_attempt = StudentQuizAttempt.objects.filter(student=rr.student, quiz=rr.quiz).order_by("-started_at").first()
    if last_attempt:
        last_attempt.retake_allowed = True
        last_attempt.save(update_fields=["retake_allowed"])
    else:
        # create a new empty attempt flagged retake_allowed so student can start
        new_attempt = StudentQuizAttempt.objects.create(
            student=rr.student,
            quiz=rr.quiz,
            started_at=timezone.now(),
            end_time=timezone.now() + timezone.timedelta(minutes=rr.quiz.duration_minutes),
            is_submitted=False,
            retake_allowed=True,
            retake_count=0,
            score=0.0
        )

    # notify student
    Notification.objects.create(sender=request.user, recipient=rr.student, message=f"Your retake request for '{rr.quiz.title}' has been approved.", role="student", is_broadcast=False)
    ActionLog.objects.create(user=request.user, action_type="approved", description="Approved retake", model_name="RetakeRequest", object_id=str(rr.id), details={"student": rr.student.username, "quiz": rr.quiz.title})

    return JsonResponse({"ok": True, "message": "Retake approved and student notified."})

